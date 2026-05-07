from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MS_BLUE = "0078D4"
WHITE = "FFFFFF"
LIGHT_GREY = "F3F2F1"
BORDER_GREY = "D1D1D1"

STATUS_FILL = {
    "GA": "DFF6DD",
    "Rolling Out": "DEECF9",
    "In Development": "FFF4CE",
    "Frontier": "FDE7F3",
}
PHASE_FILL = {
    "General Availability": "DFF6DD",
    "Targeted Release": "FFF4CE",
    "Preview": "DEECF9",
    "Frontier": "FDE7F3",
}
READINESS_FILL = {
    "Safe to Promote": "DFF6DD",
    "Pilot Only": "FFF4CE",
    "Do Not Commit": "FDE7F3",
}
CONFIDENCE_FILL = {
    "High": "DFF6DD",
    "Medium": "FFF4CE",
    "Low": "FDE7F3",
}


def _header_cell(cell, value):
    cell.value = value
    cell.fill = PatternFill(start_color=MS_BLUE, end_color=MS_BLUE, fill_type="solid")
    cell.font = Font(color=WHITE, bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border():
    side = Side(style="thin", color=BORDER_GREY)
    return Border(left=side, right=side, top=side, bottom=side)


def _color_cell(cell, hex_color):
    if hex_color:
        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def export_features_to_excel(features, custom_columns, custom_values_map):
    wb = Workbook()
    ws = wb.active
    ws.title = "Copilot Feature Matrix"
    _write_matrix(ws, features, custom_columns, custom_values_map)

    ws2 = wb.create_sheet("Summary")
    _write_summary(ws2, features)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


BASE_COLS = [
    ("Feature Name", 44),
    ("Workload", 18),
    ("Release Status", 16),
    ("Release Phase", 18),
    ("GA Estimate", 14),
    ("Confidence", 13),
    ("Business Readiness", 22),
    ("Visible in Tenant", 17),
    ("License Required", 17),
    ("Platforms", 20),
    ("Notes", 36),
    ("Source", 10),
    ("Last Updated", 14),
    ("Roadmap Link", 40),
]


def _write_matrix(ws, features, custom_columns, custom_values_map):
    all_cols = BASE_COLS + [(c.name, 18) for c in custom_columns]

    ws.row_dimensions[1].height = 36
    for idx, (col_name, col_width) in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=idx)
        _header_cell(cell, col_name)
        ws.column_dimensions[get_column_letter(idx)].width = col_width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}1"

    border = _thin_border()

    for row_idx, feat in enumerate(features, start=2):
        notes_text = " | ".join(n.note for n in feat.notes) if feat.notes else ""
        last_upd = feat.last_roadmap_update.strftime("%Y-%m-%d") if feat.last_roadmap_update else ""

        base_values = [
            feat.name,
            feat.workload or "",
            feat.release_status or "",
            feat.release_phase or "",
            feat.ga_estimate or "",
            feat.confidence_level or "",
            feat.business_readiness or "",
            feat.visible_in_tenant or "",
            feat.license_required or "",
            feat.platforms or "",
            notes_text,
            "Custom" if feat.is_custom else "MS Roadmap",
            last_upd,
            feat.roadmap_link or "",
        ]
        custom_values = [
            custom_values_map.get((feat.id, col.id), "") for col in custom_columns
        ]
        all_values = base_values + custom_values

        alt_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid") if row_idx % 2 == 0 else None

        for col_idx, value in enumerate(all_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if col_idx == 3:
                _color_cell(cell, STATUS_FILL.get(feat.release_status or ""))
            elif col_idx == 4:
                _color_cell(cell, PHASE_FILL.get(feat.release_phase or ""))
            elif col_idx == 6:
                _color_cell(cell, CONFIDENCE_FILL.get(feat.confidence_level or ""))
            elif col_idx == 7:
                _color_cell(cell, READINESS_FILL.get(feat.business_readiness or ""))
            elif alt_fill:
                cell.fill = alt_fill

        ws.row_dimensions[row_idx].height = 30


def _write_summary(ws, features):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10

    title = ws.cell(row=1, column=1, value="Copilot Feature Readiness Summary")
    title.font = Font(bold=True, size=14, color=MS_BLUE)
    ws.merge_cells("A1:C1")

    gen = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}")
    gen.font = Font(italic=True, color="666666")
    ws.merge_cells("A2:C2")

    total = len(features)

    def _section(start_row, title_text, counts_dict, fill_map):
        for c in range(1, 4):
            hdr = ws.cell(row=start_row, column=c)
            _header_cell(hdr, [title_text, "Count", "%"][c - 1])
        row = start_row + 1
        for label in sorted(counts_dict):
            count = counts_dict[label]
            pct = f"{count / total * 100:.1f}%" if total else "0%"
            a = ws.cell(row=row, column=1, value=label)
            b = ws.cell(row=row, column=2, value=count)
            cc = ws.cell(row=row, column=3, value=pct)
            hex_c = fill_map.get(label)
            if hex_c:
                for cell in (a, b, cc):
                    _color_cell(cell, hex_c)
            row += 1
        return row + 1

    status_counts = {}
    readiness_counts = {}
    phase_counts = {}
    confidence_counts = {}
    for f in features:
        status_counts[f.release_status or "Unknown"] = status_counts.get(f.release_status or "Unknown", 0) + 1
        readiness_counts[f.business_readiness or "Unknown"] = readiness_counts.get(f.business_readiness or "Unknown", 0) + 1
        phase_counts[f.release_phase or "Unknown"] = phase_counts.get(f.release_phase or "Unknown", 0) + 1
        confidence_counts[f.confidence_level or "Unknown"] = confidence_counts.get(f.confidence_level or "Unknown", 0) + 1

    next_row = _section(4, "Release Status", status_counts, STATUS_FILL)
    next_row = _section(next_row, "Release Phase", phase_counts, PHASE_FILL)
    next_row = _section(next_row, "Business Readiness", readiness_counts, READINESS_FILL)
    _section(next_row, "Confidence Level", confidence_counts, CONFIDENCE_FILL)

    total_cell = ws.cell(row=next_row + len(confidence_counts) + 2, column=1, value="TOTAL FEATURES")
    total_cell.font = Font(bold=True)
    ws.cell(row=total_cell.row, column=2, value=total).font = Font(bold=True)
